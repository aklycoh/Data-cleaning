"""读取 Excel / CSV 文件并还原合并单元格。"""

import csv
import os
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class MergeInfo:
    """描述一个合并区域。"""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def is_horizontal(self) -> bool:
        return self.max_col > self.min_col

    @property
    def is_vertical(self) -> bool:
        return self.max_row > self.min_row


@dataclass
class SheetData:
    """读取后的工作表数据。"""
    grid: list[list[Any]]
    merges: list[MergeInfo]
    n_rows: int
    n_cols: int
    row_offset: int = 0  # 实际数据起始行（跳过的前导空行数）
    col_offset: int = 0  # 实际数据起始列（跳过的前导空列数）


def _trim_grid(full_grid: list[list], return_offsets: bool = False):
    """裁剪 grid 的前导/尾部空行空列。"""
    # 前导空行
    row_offset = 0
    for row in full_grid:
        if all(v is None for v in row):
            row_offset += 1
        else:
            break

    # 前导空列
    col_offset = 0
    if full_grid:
        n_total_cols = max(len(row) for row in full_grid)
        for c in range(n_total_cols):
            if all(
                (full_grid[r][c] is None if c < len(full_grid[r]) else True)
                for r in range(len(full_grid))
            ):
                col_offset += 1
            else:
                break

    trimmed = [row[col_offset:] for row in full_grid[row_offset:]]

    # 尾部全空行
    while trimmed and all(v is None for v in trimmed[-1]):
        trimmed.pop()

    # 尾部全空列
    if trimmed:
        while trimmed[0] and all(row[-1] is None for row in trimmed):
            for row in trimmed:
                row.pop()

    n_rows = len(trimmed)
    n_cols = max((len(row) for row in trimmed), default=0)

    # 规范化：确保每行长度一致，短行用 None 填充
    for row in trimmed:
        if len(row) < n_cols:
            row.extend([None] * (n_cols - len(row)))

    if return_offsets:
        return trimmed, n_rows, n_cols, row_offset, col_offset
    return trimmed, n_rows, n_cols


def read_sheet(wb_path: str, sheet_name: str | None = None) -> SheetData:
    """读取指定工作表，还原合并单元格，返回 SheetData。"""
    wb = load_workbook(wb_path, data_only=True)
    ws: Worksheet = wb[sheet_name] if sheet_name else wb.active

    # 收集合并区域信息
    raw_merges: list[MergeInfo] = []
    for rng in ws.merged_cells.ranges:
        raw_merges.append(MergeInfo(
            min_row=rng.min_row,
            max_row=rng.max_row,
            min_col=rng.min_col,
            max_col=rng.max_col,
        ))

    # 先拆分所有合并单元格（这样才能读到左上角的值）
    merge_ranges = list(ws.merged_cells.ranges)
    for rng in merge_ranges:
        ws.unmerge_cells(str(rng))

    # 构建完整 grid
    full_grid: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column,
                            values_only=True):
        full_grid.append(list(row))

    # 填充合并区域：用左上角值填充所有被合并的单元格
    for m in raw_merges:
        val = full_grid[m.min_row - 1][m.min_col - 1]
        for r in range(m.min_row - 1, m.max_row):
            for c in range(m.min_col - 1, m.max_col):
                full_grid[r][c] = val

    wb.close()

    trimmed, n_rows, n_cols, row_offset, col_offset = _trim_grid(full_grid, return_offsets=True)

    # 调整合并区域坐标
    adjusted_merges = []
    for m in raw_merges:
        am = MergeInfo(
            min_row=m.min_row - row_offset,
            max_row=m.max_row - row_offset,
            min_col=m.min_col - col_offset,
            max_col=m.max_col - col_offset,
        )
        # 只保留在有效范围内的，并裁剪到 grid 边界
        if am.min_row >= 1 and am.min_col >= 1 and am.min_row <= n_rows and am.min_col <= n_cols:
            am.max_row = min(am.max_row, n_rows)
            am.max_col = min(am.max_col, n_cols)
            adjusted_merges.append(am)

    return SheetData(
        grid=trimmed,
        merges=adjusted_merges,
        n_rows=n_rows,
        n_cols=n_cols,
        row_offset=row_offset,
        col_offset=col_offset,
    )


def read_csv(file_path: str) -> SheetData:
    """读取 CSV 文件，返回 SheetData（无合并单元格）。"""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        full_grid = [row for row in reader]

    # 将空字符串转为 None，数值字符串转为数值
    for r, row in enumerate(full_grid):
        for c, val in enumerate(row):
            val = val.strip()
            if not val:
                full_grid[r][c] = None
            else:
                try:
                    full_grid[r][c] = float(val) if "." in val else int(val)
                except ValueError:
                    full_grid[r][c] = val

    grid, n_rows, n_cols = _trim_grid(full_grid)
    return SheetData(grid=grid, merges=[], n_rows=n_rows, n_cols=n_cols)


def read_xls(file_path: str, sheet_name: str | None = None) -> SheetData:
    """读取 .xls 文件（xlrd），还原合并单元格，返回 SheetData。"""
    import xlrd
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)

    # 构建 grid
    full_grid: list[list[Any]] = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            row.append(cell.value if cell.value != "" else None)
        full_grid.append(row)

    # 收集合并区域 (rlo, rhi, clo, chi)
    raw_merges: list[MergeInfo] = []
    for rlo, rhi, clo, chi in ws.merged_cells:
        raw_merges.append(MergeInfo(
            min_row=rlo + 1, max_row=rhi,
            min_col=clo + 1, max_col=chi,
        ))
        # 填充合并区域
        val = full_grid[rlo][clo]
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                full_grid[r][c] = val

    grid, n_rows, n_cols, row_offset, col_offset = _trim_grid(full_grid, return_offsets=True)

    adjusted_merges = []
    for m in raw_merges:
        am = MergeInfo(
            min_row=m.min_row - row_offset,
            max_row=m.max_row - row_offset,
            min_col=m.min_col - col_offset,
            max_col=m.max_col - col_offset,
        )
        if am.min_row >= 1 and am.min_col >= 1 and am.min_row <= n_rows and am.min_col <= n_cols:
            am.max_row = min(am.max_row, n_rows)
            am.max_col = min(am.max_col, n_cols)
            adjusted_merges.append(am)

    return SheetData(
        grid=grid, merges=adjusted_merges,
        n_rows=n_rows, n_cols=n_cols,
        row_offset=row_offset, col_offset=col_offset,
    )


def get_xls_sheet_names(file_path: str) -> list[str]:
    """返回 .xls 工作簿的所有工作表名称。"""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    return wb.sheet_names()


def get_sheet_names(wb_path: str) -> list[str]:
    """返回工作簿的所有工作表名称。"""
    ext = os.path.splitext(wb_path)[1].lower()
    if ext == ".xls":
        return get_xls_sheet_names(wb_path)
    wb = load_workbook(wb_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names
